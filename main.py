import sys
import sqlite3
import re
import openpyxl
from datetime import date
from PyQt6.QtSql import *
from PyQt6.QtCore import *
from PyQt6 import uic
from PyQt6.QtWidgets import *
from PyQt6 import QtGui

Form_main, Window_main = uic.loadUiType("main_form.ui")
Form_add, Window_add = uic.loadUiType("add_form.ui")
Form_edit, Window_edit = uic.loadUiType("edit_form.ui")
Form_show, Window_show = uic.loadUiType("show_form.ui")
Form_confirm_deleting, Window_confirm_deleting = uic.loadUiType(
    "confirm_deleting_form.ui"
)
Form_edit_row, Window_edit_row = uic.loadUiType("edit_row_form.ui")
Form_include_eg, Window_include_eg = uic.loadUiType("include_eg_form.ui")
Form_add_to_eg, Window_add_to_eg = uic.loadUiType("add_to_eg_form.ui")

database_name = "bd_var5.db"


def connect_db(database_name):
    con = QSqlDatabase.addDatabase("QSQLITE")
    con.setDatabaseName(database_name)
    if not con.open():
        print("Database Error: %s" % con.lastError().databaseText())
        sys.exit(1)
    print("Connection succeeded")
    # con.close()


# ---Отображение таблиц


def clicked_region():
    table_model_region = QSqlTableModel()
    table_model_region.setTable("Reg_obl_city")
    table_model_region.select()
    table_model_region.setHeaderData(0, Qt.Orientation.Horizontal, "Регион")
    table_model_region.setHeaderData(1, Qt.Orientation.Horizontal, "Область")
    table_model_region.setHeaderData(2, Qt.Orientation.Horizontal, "Город")
    # loading all data
    while table_model_region.canFetchMore():
        table_model_region.fetchMore()
    table_model_region.rowCount()
    #
    form_show.databaseTableView.setSortingEnabled(True)
    form_show.databaseTableView.setModel(table_model_region)
    form_show.databaseTableView.resizeColumnsToContents()
    form_show.databaseTableView.verticalHeader().setVisible(False)
    form_show.databaseTableView.setEditTriggers(
        QAbstractItemView.EditTrigger.NoEditTriggers
    )


def clicked_grnti():
    table_model_grnti = QSqlTableModel()
    table_model_grnti.setTable("grntirub")
    table_model_grnti.select()
    table_model_grnti.setHeaderData(0, Qt.Orientation.Horizontal, "Код")
    table_model_grnti.setHeaderData(1, Qt.Orientation.Horizontal, "Рубрика")
    # loading all data
    while table_model_grnti.canFetchMore():
        table_model_grnti.fetchMore()
    table_model_grnti.rowCount()
    #
    form_show.databaseTableView.setSortingEnabled(True)
    form_show.databaseTableView.setModel(table_model_grnti)
    form_show.databaseTableView.resizeColumnsToContents()
    form_show.databaseTableView.verticalHeader().setVisible(False)
    form_show.databaseTableView.setEditTriggers(
        QAbstractItemView.EditTrigger.NoEditTriggers
    )


# ---Добавление данных в таблицу


def get_reg_data(database_name):
    con = sqlite3.connect(database_name)
    cur = con.cursor()
    data = cur.execute("SELECT DISTINCT region FROM Reg_obl_city").fetchall()
    cur.close()
    con.close()
    return data


def populate_region_combobox():
    region_list = []
    data = get_reg_data(database_name)
    for x in data:
        region_list.append(str(x)[2:-3])
    form_add.regionComboBox.addItems(sorted(region_list))


def populate_region_filter_combobx():
    region_list = []
    data = get_reg_data(database_name)
    for x in data:
        region_list.append(str(x)[2:-3])
    form_edit.regionFilterComboBox.addItems(sorted(region_list))


def get_obl_data(database_name):
    con = sqlite3.connect(database_name)
    cur = con.cursor()
    data = cur.execute("SELECT DISTINCT oblname FROM Reg_obl_city").fetchall()
    cur.close()
    con.close()
    return data


def populate_obl_combox():
    obl_list = []
    data = get_obl_data(database_name)
    for x in data:
        obl_list.append(str(x)[2:-3])
    form_add.oblComboBox.addItems(sorted(obl_list))


def populate_obl_filter_combox():
    obl_list = []
    data = get_obl_data(database_name)
    for x in data:
        obl_list.append(str(x)[2:-3])
    form_edit.oblFilterComboBox.addItems(sorted(obl_list))


def get_city_data(database_name):
    con = sqlite3.connect(database_name)
    cur = con.cursor()
    data = cur.execute("SELECT DISTINCT city FROM Reg_obl_city").fetchall()
    cur.close()
    con.close()
    return data


def populate_city_combobox():
    city_list = []
    data = get_city_data(database_name)
    for x in data:
        city_list.append(str(x)[2:-3])
    form_add.cityComboBox.addItems(sorted(city_list))


def populate_city_filter_combobox():
    city_list = []
    data = get_city_data(database_name)
    for x in data:
        city_list.append(str(x)[2:-3])
    form_edit.cityFilterComboBox.addItems(sorted(city_list))


def id_count():
    con = sqlite3.connect(database_name)
    cur = con.cursor()
    last_id = cur.execute(
        "SELECT kod FROM Expert_final ORDER BY kod DESC LIMIT 1"
    ).fetchall()
    cur.close()
    con.close()
    return sum(last_id[0])


def get_codrub_data(database_name):
    con = sqlite3.connect(database_name)
    cur = con.cursor()
    data = cur.execute("SELECT DISTINCT codrub FROM grntirub").fetchall()
    cur.close()
    con.close()
    return data


def populate_codrub_filter_combobox():
    codrub_list = []
    data = get_codrub_data(database_name)
    for x in data:
        codrub_list.append(str(x)[2:-3])
    form_edit.grntiFilterComboBox.addItems(sorted(codrub_list))


def get_rubrika_data(database_name):
    con = sqlite3.connect(database_name)
    cur = con.cursor()
    data = cur.execute("SELECT DISTINCT rubrika FROM grntirub").fetchall()
    cur.close()
    con.close()
    return data


def populate_rubrika_filter_combobox():
    rubrika_list = []
    data = get_rubrika_data(database_name)
    for x in data:
        rubrika_list.append(str(x)[2:-3])
    form_edit.keyWordsFilterComboBox.addItems(sorted(rubrika_list))


# ---Проверки ввода


def check_name_input(name):
    name = " ".join(name.split())
    name = "".join(i for i in name if not i.isdigit())
    spaces = [i + 1 for i, j in enumerate(name) if j == " "]
    capital_letters = [i for i, j in enumerate(name) if j.isupper()]
    if len(name) == 0:
        print("Incorrect name")
        return False
    elif (
        name[-1] == "."
        and name[-3] == "."
        and name[-2].isupper()
        and name[-4].isupper()
        and name[0].isupper()
    ):
        name = "".join(name.split())
        name = name[:-4] + " " + name[-4:]
        return name
    elif name[-1] == "." and name[-2].isupper() and name[0].isupper():
        name = "".join(name.split())
        name = name[:-2] + " " + name[-2:]
        return name
    elif (
        len(name.split()) == 3
        and len([idx for idx in range(len(name)) if name[idx].isupper()]) == 3
        and spaces == capital_letters[1:3]
    ):
        return name
    else:
        print("Incorrect name")
        return False


def check_grnti_input(grnti):
    grnti = re.sub("[^0-9.]", "", grnti)
    grnti_dict = get_grntirub_dict()
    if len(grnti) == 0:
        print("Incorrect grnti")
        return False
    elif len(grnti) == 8 and grnti[:2] in grnti_dict:
        grnti1 = grnti[:8]
        grnti2 = ""
        grnti_search = grnti1[:2]
        return [grnti1, grnti2, grnti_search]
    elif len(grnti) == 16 and (grnti[:2] in grnti_dict) and (grnti[8:10] in grnti_dict):
        grnti1 = grnti[:8]
        grnti2 = grnti[8:]
        grnti_search = grnti1[:2] + " " + grnti2[:2]
        return [grnti1, grnti2, grnti_search]
    else:
        print("Incorrect grnti")
        return False


def same_person_check(name, city):
    con = sqlite3.connect(database_name)
    cur = con.cursor()
    res = cur.execute(
        """SELECT COUNT(*) FROM Expert_final WHERE name = '{}' AND city = '{}'""".format(
            name, city
        )
    ).fetchall()
    res = sum(res[0])
    cur.close()
    con.close()
    return res


def get_obl_region_dict():
    con = sqlite3.connect(database_name)
    cur = con.cursor()
    data_tuple = cur.execute("SELECT region,oblname FROM Reg_obl_city ").fetchall()
    cur.close()
    con.close()
    data_list = list(set(data_tuple))
    region_list = []
    obl_list = []
    for i in range(len(data_list)):
        region_list.append(data_list[i][0])
        obl_list.append(data_list[i][1])
    res_dict = dict(zip(obl_list, region_list))
    return res_dict


def get_city_region_dict():
    con = sqlite3.connect(database_name)
    cur = con.cursor()
    data_tuple = cur.execute("SELECT region,city FROM Reg_obl_city ").fetchall()
    cur.close()
    con.close()
    data_list = list(set(data_tuple))
    region_list = []
    city_list = []
    for i in range(len(data_list)):
        region_list.append(data_list[i][0])
        city_list.append(data_list[i][1])
    res_dict = dict(zip(city_list, region_list))
    return res_dict


def get_city_obl_dict():
    con = sqlite3.connect(database_name)
    cur = con.cursor()
    data_tuple = cur.execute("SELECT oblname,city FROM Reg_obl_city ").fetchall()
    cur.close()
    con.close()
    data_list = list(set(data_tuple))
    obl_list = []
    city_list = []
    for i in range(len(data_list)):
        obl_list.append(data_list[i][0])
        city_list.append(data_list[i][1])
    res_dict = dict(zip(city_list, obl_list))
    return res_dict


def get_obl(city):
    city_obl_dict = get_city_obl_dict()
    return city_obl_dict[city]


def region_city_check(region, city):
    res_dict = get_city_region_dict()
    return city in res_dict and region == res_dict[city]


def get_grntirub_dict():
    con = sqlite3.connect(database_name)
    cur = con.cursor()
    data_tuple = cur.execute("SELECT codrub,rubrika FROM grntirub ").fetchall()
    cur.close()
    con.close()
    data_list = list(set(data_tuple))
    codrub_list = []
    rubrika_list = []
    for i in range(len(data_list)):
        codrub_list.append(data_list[i][0])
        rubrika_list.append(data_list[i][1])
    res_dict = dict(zip(codrub_list, rubrika_list))
    return res_dict


def get_key_words(grnti_list):
    key_words_list = []
    grnti_rub_dict = get_grntirub_dict()
    for kod in grnti_list.split():
        key_words_list.append(grnti_rub_dict[kod])
        key_words_list.append("; ")
    key_words = "".join(key_words_list)
    return key_words


# ---Ввод данных


def prepare_add_form():
    form_add.nameLineAdd.clear()
    form_add.grntiLineAdd.clear()
    form_add.regionComboBox.clear()
    form_add.cityComboBox.clear()
    populate_region_combobox()


def update_add_form_city_combo_box():
    city_region_dict = get_city_region_dict()
    current_region = str(form_add.regionComboBox.currentText()).strip()
    cities_list = [k for k, v in city_region_dict.items() if v == current_region]
    if cities_list:
        form_add.cityComboBox.clear()
        form_add.cityComboBox.addItems(sorted(cities_list))


def get_input_data():
    name = str(form_add.nameLineAdd.text()).strip()
    region = str(form_add.regionComboBox.currentText()).strip()
    city = str(form_add.cityComboBox.currentText()).strip()
    grnti = str(form_add.grntiLineAdd.text()).strip()
    grnti_list = check_grnti_input(grnti)
    name = check_name_input(name)
    kod = id_count() + 1
    today = date.today()
    input_date = today.strftime("%d.%m.%Y")
    if grnti_list and name and region_city_check(region, city):
        key_words = get_key_words(grnti_list[2])
        obl = get_obl(city)
        values = [
            kod,
            name,
            region,
            city,
            grnti_list[0],
            grnti_list[1],
            key_words,
            0,
            input_date,
            grnti_list[2],
            "Не состоит",
            obl,
        ]
        if same_person_check(name, city):
            msg = QMessageBox()
            msg.setWindowIcon(QtGui.QIcon("icon.png"))
            msg.setIcon(QMessageBox.Icon.Question)
            msg.setWindowTitle("Внимание")
            msg.setText(
                "Человек с таким же ФИО в указанном городе уже находится в базе данных. Вы действительно хотите продолжить?"
            )
            msg.setStandardButtons(
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            buttonY = msg.button(QMessageBox.StandardButton.Yes)
            buttonY.setText("Да")
            buttonN = msg.button(QMessageBox.StandardButton.No)
            buttonN.setText("Нет")
            msg.exec()
            if msg.clickedButton() == buttonY:
                insert_into_db(values)
                form_add.nameLineAdd.clear()
                form_add.grntiLineAdd.clear()
                table_model.select()
                load_all_data()
                window_add.close()
            elif msg.clickedButton() == buttonN:
                msg.close()
        else:
            insert_into_db(values)
            form_add.nameLineAdd.clear()
            form_add.grntiLineAdd.clear()
            window_add.close()
            table_model.select()
            load_all_data()
    else:
        msg = QMessageBox()
        msg.setWindowIcon(QtGui.QIcon("icon.png"))
        msg.setIcon(QMessageBox.Icon.Critical)
        msg.setWindowTitle("Ошибка")
        msg.setText("Неверный ввод данных")
        msg.exec()


def insert_into_db(values):
    con = sqlite3.connect(database_name)
    cur = con.cursor()
    cur.execute(
        """INSERT INTO Expert_final VALUES(?,?,?,?,?,?,?,?,?,?,?,?)""", values
    ).fetchall()
    con.commit()
    cur.close()
    con.close()


# ---Работа с таблицами


def update_form_edit_buttons():
    form_edit.deleteDataButton.setEnabled(True)
    form_edit.exportExpertDataButton.setEnabled(True)
    form_edit.addExpertToGroupButton.setEnabled(True)


def confirm_deletion():
    msg = QMessageBox()
    msg.setWindowIcon(QtGui.QIcon("icon.png"))
    msg.setIcon(QMessageBox.Icon.Question)
    msg.setWindowTitle("Подтверждение действия")
    msg.setText("Вы действительно хотите удалить выбранные строки?")
    msg.setStandardButtons(
        QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
    )
    buttonY = msg.button(QMessageBox.StandardButton.Yes)
    buttonY.setText("Да")
    buttonN = msg.button(QMessageBox.StandardButton.No)
    buttonN.setText("Нет")
    msg.exec()
    if msg.clickedButton() == buttonY:
        delete_selected()
    elif msg.clickedButton() == buttonN:
        msg.close()


def load_all_data():
    while table_model.canFetchMore():
        table_model.fetchMore()
    table_model.rowCount()


def update_filter_obl_combo_box_by_region():
    obl_region_dict = get_obl_region_dict()
    current_region = str(form_edit.regionFilterComboBox.currentText()).strip()
    obl_list = [k for k, v in obl_region_dict.items() if v == current_region]
    if obl_list and not form_edit.cityFilterCheckBox.isChecked():
        form_edit.oblFilterComboBox.clear()
        form_edit.oblFilterComboBox.addItems(sorted(obl_list))
        form_edit.oblFilterComboBox.setCurrentIndex(-1)
    elif not form_edit.oblFilterCheckBox.isChecked():
        form_edit.oblFilterComboBox.setCurrentIndex(-1)
    elif (
        form_edit.regionFilterCheckBox.isChecked()
        and form_edit.oblFilterCheckBox.isChecked()
    ):
        form_edit.oblFilterComboBox.clear()
        form_edit.oblFilterComboBox.addItems(sorted(obl_list))
        form_edit.oblFilterComboBox.setCurrentIndex(-1)
    form_edit.databaseEditTableView.scrollToTop()


def update_filter_city_combo_box_by_region():
    city_region_dict = get_city_region_dict()
    current_region = str(form_edit.regionFilterComboBox.currentText()).strip()
    cities_list = [k for k, v in city_region_dict.items() if v == current_region]
    if cities_list and not form_edit.cityFilterCheckBox.isChecked():
        form_edit.cityFilterComboBox.clear()
        form_edit.cityFilterComboBox.addItems(sorted(cities_list))
        form_edit.cityFilterComboBox.setCurrentIndex(-1)
    elif not form_edit.cityFilterCheckBox.isChecked():
        form_edit.cityFilterComboBox.setCurrentIndex(-1)
    elif (
        form_edit.regionFilterCheckBox.isChecked()
        and form_edit.cityFilterCheckBox.isChecked()
    ):
        form_edit.cityFilterComboBox.clear()
        form_edit.cityFilterComboBox.addItems(sorted(cities_list))
        form_edit.cityFilterComboBox.setCurrentIndex(-1)
    form_edit.databaseEditTableView.scrollToTop()


def update_filter_city_combo_box_by_obl():
    city_obl_dict = get_city_obl_dict()
    current_obl = str(form_edit.oblFilterComboBox.currentText()).strip()
    cities_list = [k for k, v in city_obl_dict.items() if v == current_obl]
    if cities_list and not form_edit.cityFilterCheckBox.isChecked():
        form_edit.cityFilterComboBox.clear()
        form_edit.cityFilterComboBox.addItems(sorted(cities_list))
        form_edit.cityFilterComboBox.setCurrentIndex(-1)
    elif not form_edit.cityFilterCheckBox.isChecked():
        form_edit.cityFilterComboBox.setCurrentIndex(-1)
    elif (
        form_edit.oblFilterCheckBox.isChecked()
        and form_edit.cityFilterCheckBox.isChecked()
    ):
        form_edit.cityFilterComboBox.clear()
        form_edit.cityFilterComboBox.addItems(sorted(cities_list))
        form_edit.cityFilterComboBox.setCurrentIndex(-1)
    form_edit.databaseEditTableView.scrollToTop()


def region_check_box():
    if form_edit.regionFilterCheckBox.isChecked():
        form_edit.regionFilterComboBox.setEnabled(True)
        form_edit.regionFilterComboBox.setCurrentIndex(-1)

    else:
        form_edit.regionFilterComboBox.setEnabled(False)
        form_edit.regionFilterComboBox.setCurrentIndex(-1)
    form_edit.databaseEditTableView.scrollToTop()


def obl_check_box():
    if form_edit.oblFilterCheckBox.isChecked():
        form_edit.oblFilterComboBox.setEnabled(True)
        if not form_edit.regionFilterCheckBox.isChecked():
            form_edit.oblFilterComboBox.clear()
            populate_obl_filter_combox()
        else:
            update_filter_obl_combo_box_by_region()
        form_edit.oblFilterComboBox.setCurrentIndex(-1)

    else:
        form_edit.oblFilterComboBox.setEnabled(False)
        form_edit.oblFilterComboBox.setCurrentIndex(-1)
    form_edit.databaseEditTableView.scrollToTop()


def city_check_box():
    if form_edit.cityFilterCheckBox.isChecked():
        form_edit.cityFilterComboBox.setEnabled(True)
        if (
            not form_edit.regionFilterCheckBox.isChecked()
            and not form_edit.oblFilterCheckBox.isChecked()
        ) or (
            not form_edit.regionFilterCheckBox.isChecked()
            and form_edit.oblFilterCheckBox.isChecked()
            and form_edit.oblFilterComboBox.currentIndex() == -1
        ):
            form_edit.cityFilterComboBox.clear()
            populate_city_filter_combobox()
        elif (
            form_edit.regionFilterCheckBox.isChecked()
            and not form_edit.oblFilterCheckBox.isChecked()
        ) or (
            form_edit.regionFilterCheckBox.isChecked()
            and form_edit.oblFilterCheckBox.isChecked()
            and form_edit.oblFilterComboBox.currentIndex == -1
        ):
            update_filter_city_combo_box_by_region()
        elif (
            form_edit.regionFilterCheckBox.isChecked()
            and not form_edit.oblFilterCheckBox.isChecked()
        ) or (
            form_edit.regionFilterCheckBox.isChecked()
            and form_edit.oblFilterCheckBox.isChecked()
            and not form_edit.oblFilterComboBox.currentIndex == -1
        ):
            update_filter_city_combo_box_by_obl()
        form_edit.cityFilterComboBox.setCurrentIndex(-1)
    else:
        form_edit.cityFilterComboBox.setEnabled(False)
        form_edit.cityFilterComboBox.setCurrentIndex(-1)
    form_edit.databaseEditTableView.scrollToTop()


def grnti_radio():
    if form_edit.grntiFilterRadioButton.isChecked():
        form_edit.keyWordsFilterRadioButton.setChecked(False)
        form_edit.grntiFilterComboBox.setEnabled(True)
        form_edit.grntiFilterComboBox.setCurrentIndex(0)
    else:
        form_edit.grntiFilterComboBox.setEnabled(False)
        form_edit.grntiFilterComboBox.setCurrentIndex(-1)
    form_edit.databaseEditTableView.scrollToTop()


def key_words_radio():
    if form_edit.keyWordsFilterRadioButton.isChecked():
        form_edit.grntiFilterRadioButton.setChecked(False)
        form_edit.keyWordsFilterComboBox.setEnabled(True)
        form_edit.keyWordsFilterComboBox.setCurrentIndex(0)
    else:
        form_edit.keyWordsFilterComboBox.setEnabled(False)
        form_edit.keyWordsFilterComboBox.setCurrentIndex(-1)
    form_edit.databaseEditTableView.scrollToTop()


def get_selected_kod():
    rows_selected_kod = []
    indexes = form_edit.databaseEditTableView.selectionModel().selectedRows()
    for index in indexes:
        row = index.row()
        selected_rows = [
            proxy_model_obl.index(row, col).data()
            for col in range(proxy_model_obl.columnCount())
        ]
        rows_selected_kod.append(selected_rows[0])

    # print(rows_selected_kod)
    indexes.clear()
    return rows_selected_kod


def get_selected_data():
    indexes = form_edit.databaseEditTableView.selectionModel().selectedRows()
    for index in indexes:
        row = index.row()
        selected_rows = [
            proxy_model_obl.index(row, col).data()
            for col in range(proxy_model_obl.columnCount())
        ]
    return selected_rows


def populate_edit_row_form_region_combobox():
    form_edit_row.regionRowComboBox.clear()
    region_list = []
    data = get_reg_data(database_name)
    for x in data:
        region_list.append(str(x)[2:-3])
    form_edit_row.regionRowComboBox.addItems(sorted(region_list))


def update_edit_row_form_city_combobox():
    city_region_dict = get_city_region_dict()
    current_region = str(form_edit_row.regionRowComboBox.currentText()).strip()
    cities_list = [k for k, v in city_region_dict.items() if v == current_region]
    if cities_list:
        form_edit_row.cityRowComboBox.clear()
        form_edit_row.cityRowComboBox.addItems(sorted(cities_list))


def populate_edit_form():
    populate_edit_row_form_region_combobox()
    data = get_selected_data()
    form_edit_row.kodRowEdit.setText(str(data[0]))
    form_edit_row.kodRowEdit.setEnabled(False)
    form_edit_row.nameRowEdit.setText(str(data[1]))
    form_edit_row.regionRowComboBox.setCurrentText(str(data[2]))
    form_edit_row.cityRowComboBox.setCurrentText(str(data[3]))
    form_edit_row.grntiRowEdit.setText(str(str(data[4]) + " " + str(data[5])))
    form_edit_row.inputDateRowEdit.setText(str(data[8]))


def edit_db_row(values):
    query = QSqlQuery()
    query.prepare(
        """UPDATE Expert_final 
                            SET name = '{}',
                                region = '{}',
                                city = '{}',
                                grnti1 = '{}',
                                grnti2 = '{}',
                                key_words = '{}',
                                take_part = '{}',
                                input_date = '{}',
                                grnti_search = '{}',
                                oblname = '{}'
                            WHERE kod = {}""".format(
            values[1],
            values[2],
            values[3],
            values[4],
            values[5],
            values[6],
            values[7],
            values[8],
            values[9],
            values[10],
            values[0],
        )
    )
    res = query.exec()
    if not res:
        msg = QMessageBox()
        msg.setWindowIcon(QtGui.QIcon("icon.png"))
        msg.setIcon(QMessageBox.Icon.Critical)
        msg.setWindowTitle("Ошибка")
        msg.setText(str(query.lastError().text())+"\n"+str(query.lastError().type()))
        msg.exec()


def edit_row():
    kod = int(form_edit_row.kodRowEdit.text())
    name = str(form_edit_row.nameRowEdit.text()).strip()
    name = check_name_input(name)
    region = str(form_edit_row.regionRowComboBox.currentText()).strip()
    city = str(form_edit_row.cityRowComboBox.currentText()).strip()
    grnti = str(form_edit_row.grntiRowEdit.text()).strip()
    grnti_list = check_grnti_input(grnti)
    input_date = str(form_edit_row.inputDateRowEdit.text()).strip()
    if grnti_list and name and region_city_check(region, city):
        key_words = get_key_words(grnti_list[2])
        obl = get_obl(city)
        values = [
            kod,
            name,
            region,
            city,
            grnti_list[0],
            grnti_list[1],
            key_words,
            0,
            input_date,
            grnti_list[2],
            obl,
        ]
        edit_db_row(values)
        window_edit_row.close()
        table_model.select()
        load_all_data()
        window_edit.show()
    else:
        msg = QMessageBox()
        msg.setWindowIcon(QtGui.QIcon("icon.png"))
        msg.setIcon(QMessageBox.Icon.Critical)
        msg.setWindowTitle("Ошибка")
        msg.setText("Неверный ввод данных")
        msg.exec()


def delete_selected():
    rows_selected_kod = get_selected_kod()
    for kod in rows_selected_kod:
        query = QSqlQuery()
        query.prepare("DELETE FROM Expert_final WHERE kod=?")
        query.bindValue(0, kod)
        res = query.exec()
        if not res:
            msg = QMessageBox()
            msg.setWindowIcon(QtGui.QIcon("icon.png"))
            msg.setIcon(QMessageBox.Icon.Critical)
            msg.setWindowTitle("Ошибка")
            msg.setText(str(query.lastError().text())+"\n"+str(query.lastError().type()))
            msg.exec()
    table_model.select()
    load_all_data()


# ---Экспертная группа


def export_expert_data():
    thin_border = openpyxl.styles.borders.Border(
        left=openpyxl.styles.borders.Side(
            border_style=openpyxl.styles.borders.BORDER_THIN, color="00000000"
        ),
        right=openpyxl.styles.borders.Side(
            border_style=openpyxl.styles.borders.BORDER_THIN, color="00000000"
        ),
        top=openpyxl.styles.borders.Side(
            border_style=openpyxl.styles.borders.BORDER_THIN, color="00000000"
        ),
        bottom=openpyxl.styles.borders.Side(
            border_style=openpyxl.styles.borders.BORDER_THIN, color="00000000"
        ),
    )
    wb = openpyxl.Workbook()
    if get_selected_kod():
        kod = get_selected_kod()[0]
        con = sqlite3.connect(database_name)
        cur = con.cursor()
        cur.execute(
            """SELECT 
                        kod, 
                        name, 
                        region, 
                        oblname, 
                        city, 
                        grnti1, 
                        grnti2, 
                        key_words, 
                        take_part, 
                        input_date 
            FROM Expert_final WHERE kod = '{}'""".format(
                kod
            )
        )
        res = cur.fetchall()
        cur.close()
        con.close()
        for row in res:
            ws = wb.active
            ws.set_printer_settings(9, "landscape")
            ws.title = str(row[1])
            titles = [
                "Код",
                "ФИО",
                "Регион",
                "Область",
                "Город",
                "ГРНТИ1",
                "ГРНТИ2",
                "Ключевые слова",
                "ЧУ",
                "Дата ввода",
            ]
            for rows in ws.iter_rows(min_row=1, max_col=10, max_row=1):
                i = 0
                for cell in rows:
                    cell.font = openpyxl.styles.Font(name="Arial", size=10, bold=True)
                    cell.value = titles[i]
                    cell.border = thin_border
                    i += 1
            j = 1
            for col in row:
                cell = ws.cell(row=2, column=j)
                cell.font = openpyxl.styles.Font(name="Arial", size=10)
                cell.value = col
                cell.border = thin_border
                j += 1
        adjust_card_column_width(ws)
        wb.save("{}.xlsx".format(res[0][1]))


def get_table_names(database_name):
    con = sqlite3.connect(database_name)
    cur = con.cursor()
    data = cur.execute("SELECT name FROM sqlite_schema").fetchall()
    con.commit()
    cur.close()
    con.close()
    return data


def populate_eg_names_combobox():
    form_add_to_eg.existingExpertGroupNamesComboBox.clear()
    names_list = []
    data = get_table_names(database_name)[4:]
    for x in data:
        names_list.append(str(x)[2:-3])
    form_add_to_eg.existingExpertGroupNamesComboBox.addItems(sorted(names_list))


def populate_eg_names_to_confirm_combobox():
    form_include_eg.expertGroupComboBox.clear()
    eg_list = []
    data = get_table_names(database_name)[4:]
    for x in data:
        eg_list.append(str(x)[2:-3])
    form_include_eg.expertGroupComboBox.addItems(sorted(eg_list))
    form_include_eg.expertGroupComboBox.setCurrentIndex(-1)


def new_group_radio():
    form_add_to_eg.expertGroupNameLineEdit.setEnabled(True)
    form_add_to_eg.existingExpertGroupNamesComboBox.setEnabled(False)


def select_group_radio():
    form_add_to_eg.expertGroupNameLineEdit.setEnabled(False)
    form_add_to_eg.existingExpertGroupNamesComboBox.setEnabled(True)


def get_expert_group_name():
    if (
        form_add_to_eg.newGroupRadioButton.isChecked()
        and form_add_to_eg.expertGroupNameLineEdit.text()
    ):
        return ["N", str(form_add_to_eg.expertGroupNameLineEdit.text()).strip()]
    if (
        form_add_to_eg.selectGroupRadioButton.isChecked()
        and form_add_to_eg.existingExpertGroupNamesComboBox.currentText()
    ):
        return [
            "E",
            str(form_add_to_eg.existingExpertGroupNamesComboBox.currentText()).strip(),
        ]
    else:
        msg = QMessageBox()
        msg.setWindowIcon(QtGui.QIcon("icon.png"))
        msg.setIcon(QMessageBox.Icon.Critical)
        msg.setWindowTitle("Ошибка")
        msg.setText("Неверный ввод данных")
        msg.exec()
        return False


def include_in_eg():
    rows_selected_kod = get_selected_kod()
    expert_group_name = get_expert_group_name()
    if expert_group_name and expert_group_name[0] == "N":
        query = QSqlQuery()
        query.prepare(
            """CREATE TABLE '{}' AS SELECT * FROM Expert_final WHERE 0""".format(
                expert_group_name[1]
            )
        )
        res = query.exec()
        if not res:
            msg = QMessageBox()
            msg.setWindowIcon(QtGui.QIcon("icon.png"))
            msg.setIcon(QMessageBox.Icon.Critical)
            msg.setWindowTitle("Ошибка")
            msg.setText(str(query.lastError().text())+"\n"+str(query.lastError().type()))
            msg.exec()

        for kod in rows_selected_kod:
            query = QSqlQuery()
            query.prepare(
                """UPDATE Expert_final SET status = 'На рассмотрении' WHERE kod='{}'""".format(
                    kod
                )
            )
            res = query.exec()
            if not res:
                msg = QMessageBox()
                msg.setWindowIcon(QtGui.QIcon("icon.png"))
                msg.setIcon(QMessageBox.Icon.Critical)
                msg.setWindowTitle("Ошибка")
                msg.setText(str(query.lastError().text())+"\n"+str(query.lastError().type()))
                msg.exec()
            query.prepare(
                """INSERT INTO '{}' SELECT * FROM Expert_final WHERE kod='{}'""".format(
                    expert_group_name[1], kod
                )
            )
            res = query.exec()
            if not res:
                msg = QMessageBox()
                msg.setWindowIcon(QtGui.QIcon("icon.png"))
                msg.setIcon(QMessageBox.Icon.Critical)
                msg.setWindowTitle("Ошибка")
                msg.setText(str(query.lastError().text())+"\n"+str(query.lastError().type()))
                msg.exec()
        table_model.select()
        load_all_data()
        delete_duplicates_in_eg(expert_group_name[1])
        table_model_eg.select()
        form_include_eg.expertGroupTableView.resizeColumnsToContents()
    if expert_group_name and expert_group_name[0] == "E":
        for kod in rows_selected_kod:
            query = QSqlQuery()
            query.prepare(
                """UPDATE Expert_final SET status = 'На рассмотрении' WHERE kod='{}'""".format(
                    kod
                )
            )
            res = query.exec()
            if not res:
                msg = QMessageBox()
                msg.setWindowIcon(QtGui.QIcon("icon.png"))
                msg.setIcon(QMessageBox.Icon.Critical)
                msg.setWindowTitle("Ошибка")
                msg.setText(str(query.lastError().text())+"\n"+str(query.lastError().type()))
                msg.exec()
            query.prepare(
                """INSERT INTO '{}' SELECT * FROM Expert_final WHERE kod='{}'""".format(
                    expert_group_name[1], kod
                )
            )
            res = query.exec()
            if not res:
                msg = QMessageBox()
                msg.setWindowIcon(QtGui.QIcon("icon.png"))
                msg.setIcon(QMessageBox.Icon.Critical)
                msg.setWindowTitle("Ошибка")
                msg.setText(str(query.lastError().text())+"\n"+str(query.lastError().type()))
                msg.exec()
        table_model.select()
        load_all_data()
        delete_duplicates_in_eg(expert_group_name[1])
        table_model_eg.select()
        form_include_eg.expertGroupTableView.resizeColumnsToContents()
    form_add_to_eg.expertGroupNameLineEdit.clear()


def delete_duplicates_in_eg(expert_group_name):
    query = QSqlQuery(
        "DELETE FROM '{}' WHERE rowid NOT IN (SELECT MIN(rowid) FROM '{}' GROUP BY kod)".format(
            expert_group_name, expert_group_name
        )
    )
    res = query.exec()
    if not res:
        msg = QMessageBox()
        msg.setWindowIcon(QtGui.QIcon("icon.png"))
        msg.setIcon(QMessageBox.Icon.Critical)
        msg.setWindowTitle("Ошибка")
        msg.setText(str(query.lastError().text())+"\n"+str(query.lastError().type()))
        msg.exec()


def delete_selected_eg():
    expert_group_name = form_include_eg.expertGroupComboBox.currentText()
    rows_selected_kod = []
    indexes = form_include_eg.expertGroupTableView.selectionModel().selectedRows()
    for index in indexes:
        rows_selected_kod.append(index.data())
    for kod in rows_selected_kod:
        query = QSqlQuery()
        query.prepare("UPDATE Expert_final SET status = 'Не состоит' WHERE kod=?")
        query.bindValue(0, kod)
        res = query.exec()
        if not res:
            msg = QMessageBox()
            msg.setWindowIcon(QtGui.QIcon("icon.png"))
            msg.setIcon(QMessageBox.Icon.Critical)
            msg.setWindowTitle("Ошибка")
            msg.setText(str(query.lastError().text())+"\n"+str(query.lastError().type()))
            msg.exec()
        query.prepare("DELETE FROM '{}' WHERE kod=?".format(expert_group_name))
        query.bindValue(0, kod)
        res = query.exec()
        if not res:
            msg = QMessageBox()
            msg.setWindowIcon(QtGui.QIcon("icon.png"))
            msg.setIcon(QMessageBox.Icon.Critical)
            msg.setWindowTitle("Ошибка")
            msg.setText(str(query.lastError().text())+"\n"+str(query.lastError().type()))
            msg.exec()
    table_model_eg.select()
    table_model.select()
    load_all_data()


def update_table_model_eg():
    table = form_include_eg.expertGroupComboBox.currentText()
    table_model_eg.setTable(table)
    table_model_eg.select()
    table_model_eg.setHeaderData(0, Qt.Orientation.Horizontal, "Код")
    table_model_eg.setHeaderData(1, Qt.Orientation.Horizontal, "ФИО")
    table_model_eg.setHeaderData(2, Qt.Orientation.Horizontal, "Регион")
    table_model_eg.setHeaderData(3, Qt.Orientation.Horizontal, "Город")
    table_model_eg.setHeaderData(4, Qt.Orientation.Horizontal, "ГРНТИ1")
    table_model_eg.setHeaderData(5, Qt.Orientation.Horizontal, "ГРНТИ2")
    table_model_eg.setHeaderData(7, Qt.Orientation.Horizontal, "Участия")
    table_model_eg.setHeaderData(8, Qt.Orientation.Horizontal, "Дата ввода")
    table_model_eg.setHeaderData(10, Qt.Orientation.Horizontal, "Статус в ЭГ")
    form_include_eg.expertGroupTableView.setModel(table_model_eg)
    form_include_eg.expertGroupTableView.setSortingEnabled(True)
    form_include_eg.expertGroupTableView.sortByColumn(0, Qt.SortOrder.AscendingOrder)
    form_include_eg.expertGroupTableView.resizeColumnsToContents()
    form_include_eg.expertGroupTableView.verticalHeader().setVisible(False)
    form_include_eg.expertGroupTableView.hideColumn(6)
    form_include_eg.expertGroupTableView.hideColumn(9)
    form_include_eg.expertGroupTableView.hideColumn(11)
    form_include_eg.expertGroupTableView.setEditTriggers(
        QAbstractItemView.EditTrigger.NoEditTriggers
    )
    form_include_eg.expertGroupTableView.setSelectionBehavior(
        QAbstractItemView.SelectionBehavior.SelectRows
    )
    form_include_eg.removeExpertButton.clicked.connect(delete_selected_eg)


def confirm_eg():
    expert_group_name = form_include_eg.expertGroupComboBox.currentText()
    if expert_group_name:
        kod_list = []
        con = sqlite3.connect(database_name)
        cur = con.cursor()
        res = cur.execute("SELECT kod FROM '{}'".format(expert_group_name)).fetchall()
        con.commit()
        cur.close()
        con.close()
        for i in range(len(res)):
            kod_list.append(res[i][0])
        # print(kod_list)
        if len(kod_list):
            for kod in kod_list:
                con = sqlite3.connect(database_name)
                cur = con.cursor()
                cur.execute(
                    "UPDATE Expert_final SET take_part = take_part + 1, status = 'Утвержден' WHERE kod = '{}'".format(
                        kod
                    )
                ).fetchall()
                con.commit()
                cur.close()
                con.close()
            if expert_group_name:
                export_to_xlsx(expert_group_name)
                con = sqlite3.connect(database_name)
                cur = con.cursor()
                cur.execute("DROP TABLE '{}'".format(expert_group_name))
                con.commit()
                cur.close()
                con.close()
            table_model_eg.select()
            table_model.select()
            load_all_data()


def export_to_xlsx(expert_group_name):
    wb = openpyxl.Workbook()
    expert_group_sheet(wb, expert_group_name)
    expert_card_sheets(wb, expert_group_name)
    wb.save("{}.xlsx".format(expert_group_name))


def expert_group_sheet(wb, expert_group_name):
    thin_border = openpyxl.styles.borders.Border(
        left=openpyxl.styles.borders.Side(
            border_style=openpyxl.styles.borders.BORDER_THIN, color="00000000"
        ),
        right=openpyxl.styles.borders.Side(
            border_style=openpyxl.styles.borders.BORDER_THIN, color="00000000"
        ),
        top=openpyxl.styles.borders.Side(
            border_style=openpyxl.styles.borders.BORDER_THIN, color="00000000"
        ),
        bottom=openpyxl.styles.borders.Side(
            border_style=openpyxl.styles.borders.BORDER_THIN, color="00000000"
        ),
    )
    ws = wb.active
    ws.title = expert_group_name
    ws.set_printer_settings(9, "landscape")
    name_cell = ws.cell(row=1, column = 1)
    ws.merge_cells('A1:F1')

    for row in ws.iter_rows(min_row=1, max_col=6, max_row=1):
        i = 0
        for cell in row:
            cell.border = thin_border
            cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
            i += 1
    name_cell.font = openpyxl.styles.Font(name="Arial", size=10, bold=True)
    name_cell.value = "Экспертная группа '{}'".format(expert_group_name)
    name_cell.alignment = openpyxl.styles.Alignment(horizontal='center')
    titles = ["№", "ФИО", "Регион", "Город", "ГРНТИ1", "ГРНТИ2"]
    for row in ws.iter_rows(min_row=2, max_col=6, max_row=2):
        i = 0
        for cell in row:
            cell.font = openpyxl.styles.Font(name="Arial", size=10, bold=True)
            cell.value = titles[i]
            cell.border = thin_border

            i += 1
    con = sqlite3.connect(database_name)
    cur = con.cursor()
    cur.execute(
        "SELECT kod, name, region, city, grnti1, grnti2 FROM '{}'".format(
            expert_group_name
        )
    )
    res = cur.fetchall()
    cur.close()
    con.close()
    i = 2
    for row in res:
        i += 1
        j = 1
        for col in row:
            cell = ws.cell(row=i, column=j)
            cell.font = openpyxl.styles.Font(name="Arial", size=10)
            cell.border = thin_border
            if j == 1:
                cell.value = i - 2
            else:
                cell.value = col
            j += 1
    adjust_group_column_width(ws)


def expert_card_sheets(wb, expert_group_name):
    thin_border = openpyxl.styles.borders.Border(
        left=openpyxl.styles.borders.Side(
            border_style=openpyxl.styles.borders.BORDER_THIN, color="00000000"
        ),
        right=openpyxl.styles.borders.Side(
            border_style=openpyxl.styles.borders.BORDER_THIN, color="00000000"
        ),
        top=openpyxl.styles.borders.Side(
            border_style=openpyxl.styles.borders.BORDER_THIN, color="00000000"
        ),
        bottom=openpyxl.styles.borders.Side(
            border_style=openpyxl.styles.borders.BORDER_THIN, color="00000000"
        ),
    )
    con = sqlite3.connect(database_name)
    cur = con.cursor()
    cur.execute(
        "SELECT kod, name, region, oblname, city, grnti1, grnti2, key_words, take_part, input_date FROM '{}'".format(
            expert_group_name
        )
    )
    res = cur.fetchall()
    cur.close()
    con.close()
    for row in res:
        ws = wb.create_sheet(str(row[1]))
        ws.set_printer_settings(9, "landscape")
        titles = [
            "Код",
            "ФИО",
            "Регион",
            "Область",
            "Город",
            "ГРНТИ1",
            "ГРНТИ2",
            "Ключевые слова",
            "ЧУ",
            "Дата ввода",
        ]
        for rows in ws.iter_rows(min_row=1, max_col=10, max_row=1):
            i = 0
            for cell in rows:
                cell.font = openpyxl.styles.Font(name="Arial", size=10, bold=True)
                cell.value = titles[i]
                cell.border = thin_border
                i += 1
        j = 1
        for col in row:
            cell = ws.cell(row=2, column=j)
            cell.font = openpyxl.styles.Font(name="Arial", size=10)
            cell.value = col
            cell.border = thin_border
            j += 1
        adjust_card_column_width(ws)


def adjust_column_width(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width


def adjust_group_column_width(ws):
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15


def adjust_card_column_width(ws):
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 16
    ws["D2"].alignment = openpyxl.styles.Alignment(wrap_text=True)
    ws.column_dimensions["E"].width = 21
    ws.column_dimensions["F"].width = 8
    ws.column_dimensions["G"].width = 8
    ws.column_dimensions["H"].width = 20
    ws["H2"].alignment = openpyxl.styles.Alignment(wrap_text=True)
    ws.column_dimensions["I"].width = 3
    ws.column_dimensions["J"].width = 11


# ---Переходы между окнами


def open_show_window():
    window_main.close()
    window_show.show()


def return_to_main_from_show():
    window_show.close()
    window_main.show()


def open_add_window():
    window_add.show()


def return_to_main_from_add():
    form_add.nameLineAdd.clear()
    form_add.grntiLineAdd.clear()
    window_add.close()


def open_edit_window():
    window_main.close()
    window_edit.show()


def return_to_main_from_edit():
    window_edit.close()
    window_main.show()


def open_edit_row_window():
    window_edit_row.show()


def return_to_edit_from_row():
    window_edit_row.close()
    window_edit.show()


def open_include_window():
    populate_eg_names_to_confirm_combobox()
    window_include_eg.show()


def return_to_edit_from_confirm_eg():
    window_include_eg.close()


def open_add_to_eg_window():
    populate_eg_names_combobox()
    window_add_to_eg.show()


def exit_main():
    app.closeAllWindows()


app = QApplication([])
window_main = Window_main()
form_main = Form_main()
form_main.setupUi(window_main)
connect_db(database_name)
window_main.setWindowIcon(QtGui.QIcon("icon.png"))
window_main.setWindowTitle("Главное меню")
# form.databaseConnectButton.clicked.connect(lambda: connect_db(database_name))
form_main.showDataButton.clicked.connect(open_show_window)
form_main.exitButton.clicked.connect(exit_main)

window_show = Window_show()
form_show = Form_show()
form_show.setupUi(window_show)
window_show.setWindowIcon(QtGui.QIcon("icon.png"))
window_show.setWindowTitle("Отображение таблиц")

form_show.GrntiTableButton.clicked.connect(clicked_grnti)
form_show.RegionTableButton.clicked.connect(clicked_region)
form_show.returnToMainButton.clicked.connect(return_to_main_from_show)


form_main.editDataButton.clicked.connect(open_edit_window)


window_edit = Window_edit()
form_edit = Form_edit()
form_edit.setupUi(window_edit)
window_edit.setWindowIcon(QtGui.QIcon("icon.png"))
window_edit.setWindowTitle("Работа с таблицами")

form_edit.deleteDataButton.setEnabled(False)
form_edit.exportExpertDataButton.setEnabled(False)
form_edit.addExpertToGroupButton.setEnabled(False)
form_edit.databaseEditTableView.clicked.connect(update_form_edit_buttons)
form_edit.addDataButton.clicked.connect(open_add_window)

window_add = Window_add()
form_add = Form_add()
form_add.setupUi(window_add)
window_add.setWindowIcon(QtGui.QIcon("icon.png"))
window_add.setWindowTitle("Добавление информации")

form_add.regionComboBox.currentTextChanged.connect(update_add_form_city_combo_box)
form_edit.addDataButton.clicked.connect(prepare_add_form)
form_add.addDataButton.clicked.connect(get_input_data)
form_add.returnToMainButton.clicked.connect(return_to_main_from_add)


table_model = QSqlTableModel()
table_model.setTable("Expert_final")
table_model.select()
table_model.setHeaderData(0, Qt.Orientation.Horizontal, "Код")
table_model.setHeaderData(1, Qt.Orientation.Horizontal, "ФИО")
table_model.setHeaderData(2, Qt.Orientation.Horizontal, "Регион")
table_model.setHeaderData(3, Qt.Orientation.Horizontal, "Город")
table_model.setHeaderData(4, Qt.Orientation.Horizontal, "ГРНТИ1")
table_model.setHeaderData(5, Qt.Orientation.Horizontal, "ГРНТИ2")
table_model.setHeaderData(7, Qt.Orientation.Horizontal, "Участия")
table_model.setHeaderData(8, Qt.Orientation.Horizontal, "Дата ввода")
table_model.setHeaderData(10, Qt.Orientation.Horizontal, "Статус в ЭГ")
load_all_data()
form_edit.databaseEditTableView.setSortingEnabled(True)
form_edit.databaseEditTableView.setModel(table_model)
form_edit.databaseEditTableView.verticalHeader().setVisible(False)
form_edit.databaseEditTableView.hideColumn(6)
# form_edit.databaseEditTableView.hideColumn(7)
form_edit.databaseEditTableView.hideColumn(9)
form_edit.databaseEditTableView.hideColumn(11)
form_edit.databaseEditTableView.setEditTriggers(
    QAbstractItemView.EditTrigger.NoEditTriggers
)
form_edit.databaseEditTableView.setSelectionBehavior(
    QAbstractItemView.SelectionBehavior.SelectRows
)


source_model = table_model

proxy_model_kod = QSortFilterProxyModel()
proxy_model_kod.setSourceModel(source_model)
form_edit.databaseEditTableView.setModel(proxy_model_kod)
proxy_model_kod.setFilterKeyColumn(0)
proxy_model_kod.setFilterRegularExpression(
    QRegularExpression(form_edit.idFilterEdit.text())
)
proxy_model_kod.setFilterCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
form_edit.idFilterEdit.textChanged.connect(proxy_model_kod.setFilterRegularExpression)

proxy_model_name = QSortFilterProxyModel()
proxy_model_name.setSourceModel(proxy_model_kod)
form_edit.databaseEditTableView.setModel(proxy_model_name)
proxy_model_name.setFilterKeyColumn(1)
proxy_model_name.setFilterRegularExpression(
    QRegularExpression(form_edit.nameFilterEdit.text())
)
proxy_model_name.setFilterCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
form_edit.nameFilterEdit.textChanged.connect(
    proxy_model_name.setFilterRegularExpression
)

populate_region_filter_combobx()
form_edit.regionFilterComboBox.setEnabled(False)
form_edit.regionFilterComboBox.setCurrentIndex(-1)
form_edit.regionFilterComboBox.currentTextChanged.connect(
    update_filter_city_combo_box_by_region
)
form_edit.regionFilterComboBox.currentTextChanged.connect(
    update_filter_obl_combo_box_by_region
)
form_edit.cityFilterComboBox.currentTextChanged.connect(
    form_edit.databaseEditTableView.scrollToTop
)
form_edit.regionFilterCheckBox.stateChanged.connect(region_check_box)


proxy_model_region = QSortFilterProxyModel()
proxy_model_region.setSourceModel(proxy_model_name)
form_edit.databaseEditTableView.setModel(proxy_model_region)
proxy_model_region.setFilterKeyColumn(2)
form_edit.regionFilterComboBox.currentTextChanged.connect(
    proxy_model_region.setFilterFixedString
)

populate_city_filter_combobox()
form_edit.cityFilterComboBox.setEnabled(False)
form_edit.cityFilterComboBox.setCurrentIndex(-1)
form_edit.cityFilterCheckBox.stateChanged.connect(city_check_box)
proxy_model_city = QSortFilterProxyModel()
proxy_model_city.setSourceModel(proxy_model_region)
form_edit.databaseEditTableView.setModel(proxy_model_city)
proxy_model_city.setFilterKeyColumn(3)
form_edit.cityFilterComboBox.currentTextChanged.connect(
    proxy_model_city.setFilterFixedString
)


populate_codrub_filter_combobox()
form_edit.grntiFilterComboBox.setEnabled(False)
form_edit.grntiFilterComboBox.setCurrentIndex(-1)
# form_edit.grntiFilterCheckBox.stateChanged.connect(grnti_check_box)
form_edit.grntiFilterRadioButton.toggled.connect(grnti_radio)
proxy_model_grnti = QSortFilterProxyModel()
proxy_model_grnti.setSourceModel(proxy_model_city)
form_edit.databaseEditTableView.setModel(proxy_model_grnti)
proxy_model_grnti.setFilterKeyColumn(9)
form_edit.grntiFilterComboBox.currentTextChanged.connect(
    proxy_model_grnti.setFilterFixedString
)

proxy_model_input_date = QSortFilterProxyModel()
proxy_model_input_date.setSourceModel(proxy_model_grnti)
form_edit.databaseEditTableView.setModel(proxy_model_input_date)
proxy_model_input_date.setFilterKeyColumn(8)
proxy_model_input_date.setFilterRegularExpression(
    QRegularExpression(form_edit.inputDateFilterEdit.text())
)
proxy_model_input_date.setFilterCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
form_edit.inputDateFilterEdit.textChanged.connect(
    proxy_model_input_date.setFilterRegularExpression
)

populate_rubrika_filter_combobox()
form_edit.keyWordsFilterComboBox.setEnabled(False)
form_edit.keyWordsFilterComboBox.setCurrentIndex(-1)
# form_edit.keyWordsFilterCheckBox.stateChanged.connect(key_words_check_box)
form_edit.keyWordsFilterRadioButton.toggled.connect(key_words_radio)
proxy_model_key_words = QSortFilterProxyModel()
proxy_model_key_words.setSourceModel(proxy_model_input_date)
form_edit.databaseEditTableView.setModel(proxy_model_key_words)
proxy_model_key_words.setFilterKeyColumn(6)
form_edit.keyWordsFilterComboBox.currentTextChanged.connect(
    proxy_model_key_words.setFilterFixedString
)

populate_obl_filter_combox()
form_edit.oblFilterComboBox.setEnabled(False)
form_edit.oblFilterComboBox.setCurrentIndex(-1)
form_edit.oblFilterCheckBox.stateChanged.connect(obl_check_box)
form_edit.oblFilterComboBox.currentTextChanged.connect(
    update_filter_city_combo_box_by_obl
)
proxy_model_obl = QSortFilterProxyModel()
proxy_model_obl.setSourceModel(proxy_model_key_words)
form_edit.databaseEditTableView.setModel(proxy_model_obl)
proxy_model_obl.setFilterKeyColumn(11)
form_edit.oblFilterComboBox.currentTextChanged.connect(
    proxy_model_obl.setFilterFixedString
)

form_edit.databaseEditTableView.resizeColumnsToContents()
form_edit.exportExpertDataButton.clicked.connect(export_expert_data)

window_edit_row = Window_edit_row()
form_edit_row = Form_edit_row()
form_edit_row.setupUi(window_edit_row)
window_edit_row.setWindowIcon(QtGui.QIcon("icon.png"))
window_edit_row.setWindowTitle("Редактирование информации")

window_add_to_eg = Window_add_to_eg()
form_add_to_eg = Form_add_to_eg()
form_add_to_eg.setupUi(window_add_to_eg)
window_add_to_eg.setWindowIcon(QtGui.QIcon("icon.png"))
window_add_to_eg.setWindowTitle("Добавление эксперта в группу")
form_add_to_eg.buttonBox.addButton("Отмена", QDialogButtonBox.ButtonRole.RejectRole)
form_add_to_eg.existingExpertGroupNamesComboBox.setEnabled(False)
form_add_to_eg.newGroupRadioButton.toggled.connect(new_group_radio)
form_add_to_eg.selectGroupRadioButton.toggled.connect(select_group_radio)
window_include_eg = Window_include_eg()
form_include_eg = Form_include_eg()
form_include_eg.setupUi(window_include_eg)
window_include_eg.setWindowIcon(QtGui.QIcon("icon.png"))
window_include_eg.setWindowTitle("Утверждение экспертной группы")

form_edit.databaseEditTableView.doubleClicked.connect(populate_edit_form)
form_edit.databaseEditTableView.doubleClicked.connect(open_edit_row_window)
form_edit_row.saveChangesButton.clicked.connect(edit_row)
form_edit_row.regionRowComboBox.currentTextChanged.connect(
    update_edit_row_form_city_combobox
)
form_edit_row.cancelButton.clicked.connect(return_to_edit_from_row)
form_edit.deleteDataButton.clicked.connect(confirm_deletion)
form_edit.returnToMainButton.clicked.connect(return_to_main_from_edit)


form_include_eg.returnToEditButton.clicked.connect(return_to_edit_from_confirm_eg)
table_model_eg = QSqlTableModel()
form_include_eg.expertGroupComboBox.currentTextChanged.connect(update_table_model_eg)

form_include_eg.confirmExpertGroupButton.clicked.connect(confirm_eg)
form_include_eg.confirmExpertGroupButton.clicked.connect(populate_eg_names_to_confirm_combobox)

form_edit.confirmExpertGroupButton.clicked.connect(open_include_window)
form_edit.addExpertToGroupButton.clicked.connect(open_add_to_eg_window)
form_add_to_eg.buttonBox.accepted.connect(include_in_eg)


window_main.show()
app.exec()
